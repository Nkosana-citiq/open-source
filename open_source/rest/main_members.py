from datetime import datetime
import os
import csv
import uuid
import falcon
import json
import logging
import pandas as pd

from dateutil.parser import parse
from dateutil.relativedelta import relativedelta

from open_source import config, db

from open_source.core.applicants import Applicant
from open_source.core.consultants import Consultant
from open_source.core.plans import Plan
from sqlalchemy.orm.exc import MultipleResultsFound, NoResultFound
from sqlalchemy import or_
from open_source.core.main_members import MainMember
from open_source.core.extended_members import ExtendedMember
from open_source.rest import extended_members
from open_source.rest.extended_members import update_certificate, bulk_insert_extended_members
from open_source.core.parlours import Parlour
from open_source.utils import localize_contact

from falcon_cors import CORS
from za_id_number.za_id_number import SouthAfricanIdentityValidate

logger = logging.getLogger(__name__)
public_cors = CORS(allow_all_origins=True)

conf = config.get_config()

class MainMemberGetEndpoint:
    cors = public_cors
    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def on_get(self, req, resp, id):
        with db.transaction() as session:
            main_member = session.query(MainMember).filter(
                MainMember.id == id
            ).first()

            if main_member is None:
                raise falcon.HTTPNotFound(title="Not Found", description="Applicant Not Found")

            resp.body = json.dumps(main_member.to_dict(), default=str)


class MainMemberCountEndpoint:
    cors = public_cors

    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def on_get(self, req, resp, id):
        try:
            with db.no_transaction() as session:
                try:
                    parlour = session.query(Parlour).filter(
                        Parlour.state == Parlour.STATE_ACTIVE,
                        Parlour.id == id
                    ).one_or_none()

                except MultipleResultsFound as e:
                    raise falcon.HTTPBadRequest(title="Error", description="Error getting applicants")

                applicants = session.query(Applicant).filter(
                    Applicant.state == Applicant.STATE_ACTIVE,
                    Applicant.parlour_id == parlour.id
                ).order_by(Applicant.id.desc())

                applicant_ids = [applicant.id for applicant in applicants.all()]
                main_member_count = session.query(MainMember).filter(
                    MainMember.state == MainMember.STATE_ACTIVE,
                    MainMember.applicant_id.in_(applicant_ids)
                ).count()

                resp.body = json.dumps({"count": main_member_count}, default=str)

        except:
            logger.exception("Error, Failed to get Applicants for user with ID {}.".format(id))
            raise falcon.HTTPUnprocessableEntity(title="Uprocessable entity", description="Failed to get Applicants for user with ID {}.".format(id))


class MainGetAllParlourEndpoint:
    cors = public_cors

    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def on_get(self, req, resp, id):
        try:
            with db.transaction() as session:
                try:
                    status = None
                    search_field = None
                    search_date = None
                    consultant = None
                    consultants = None
                    start_date = None
                    end_date = None

                    if "status" in req.params:
                        status = req.params.pop("status")

                    if "search_string" in req.params:
                        search_field = req.params.pop("search_string")

                    if "search_date" in req.params:
                        search_date = req.params.pop("search_date")

                    if "start_date" in req.params:
                        start_date = req.params.pop("start_date")

                    if "end_date" in req.params:
                        end_date = req.params.pop("end_date")

                    if "consultant" in req.params:
                        consultant_id = req.params.pop("consultant")
                        consultant = session.query(Consultant).get(consultant_id)

                    if "branch" in req.params:
                        parlour_branch = req.params.pop("branch")
                        consultants = session.query(Consultant).filter(Consultant.branch == parlour_branch.strip()).all()

                    parlour = session.query(Parlour).filter(
                        Parlour.state == Parlour.STATE_ACTIVE,
                        Parlour.id == id
                    ).one_or_none()

                except MultipleResultsFound as e:
                    raise falcon.HTTPBadRequest(title="Error", description="Error getting applicants")

                if search_date:
                    search = parse(search_date)

                    main_members = session.query(MainMember).filter(MainMember.state == MainMember.STATE_ACTIVE, MainMember.parlour_id == parlour.id).all()
                    main_count = len(main_members)
                    main_members = [m for m in main_members if m.date_joined and m.date_joined.year == search.date().year and m.date_joined.month == search.date().month]
                    month_count = len(main_members)

                    resp.body = json.dumps({"original": main_count, "month": month_count, "period": '-'.join([str(search.date().year), str(search.date().month)])}, default=str)
                elif search_field:
                    main_members = session.query(
                        MainMember,
                        Applicant
                    ).join(Applicant, (MainMember.applicant_id==Applicant.id)).filter(
                        MainMember.state == MainMember.STATE_ACTIVE,
                        MainMember.parlour_id == parlour.id,
                        or_(
                            MainMember.first_name.ilike('{}%'.format(search_field)),
                            MainMember.last_name.ilike('{}%'.format(search_field)),
                            MainMember.id_number.ilike('{}%'.format(search_field)),
                            Applicant.policy_num.ilike('{}%'.format(search_field))
                        )
                    )

                    if not main_members:
                        resp.body = json.dumps({"total": 0, "count": 0, "offset": 0, "limit":0, "result": []})
                    else:
                        result = MainMember._paginated_search_results(req.params, main_members)
                        resp.body = json.dumps(result, default=str)
                else:
                    applicants = session.query(Applicant).filter(
                        Applicant.state == Applicant.STATE_ACTIVE,
                        Applicant.parlour_id == parlour.id
                    ).order_by(Applicant.id.desc())

                    if consultant:
                        applicants = applicants.filter(Applicant.consultant_id == consultant.id)

                    if consultants:
                        consultant_ids = [consultant.id for consultant in consultants]
                        applicants = applicants.filter(Applicant.consultant_id.in_(consultant_ids))

                    if status:
                        applicants = applicants.filter(Applicant.status == status.lower())

                    applicant_res = [(applicant, applicant.plan) for applicant in applicants.limit(100).all()]
                    if applicant_res:
                        for applicant in applicant_res:
                            if applicant[0].plan.id == applicant[1].id:
                                max_age_limit = applicant[1].member_maximum_age
                                min_age_limit = applicant[1].member_minimum_age
                                main_member = session.query(MainMember).filter(
                                    MainMember.state == MainMember.STATE_ACTIVE,
                                    MainMember.applicant_id == applicant[0].id
                                ).first()
                                if main_member:
                                    id_number = main_member.id_number
                                    number = int(id_number[0:2])
                                    if number == 0:
                                        number = 2000
                                    try:
                                        dob = datetime(number, int(id_number[2:4]), int(id_number[4:6]),0,0,0,0)
                                    except:
                                        dob = datetime(number, int(id_number[2:4]),1,0,0,0,0)
                                    now = datetime.now()
                                    age = relativedelta(now, dob)

                                    years = str(age.years)[2:] if str(age.years)[2:].isdigit() else str(age.years)
                                    if max_age_limit and int(years) > max_age_limit:
                                        main_member.age_limit_exceeded = True
                                    if min_age_limit and int(years) < min_age_limit:
                                        main_member.age_limit_exceeded = True
                                    session.commit()
                    applicant_ids = [applicant.id for applicant in applicants.all()]
                    main_members = session.query(MainMember).filter(
                        MainMember.state == MainMember.STATE_ACTIVE,
                        MainMember.applicant_id.in_(applicant_ids)
                    ).order_by(MainMember.id.desc())

                    if start_date:
                        main_members = main_members.filter(
                            MainMember.created_at >= start_date
                        )

                    if end_date:
                        main_members = main_members.filter(
                            MainMember.created_at <= end_date
                        )
                    if not main_members.all():
                        resp.body = json.dumps({})
                    else:
                        result = MainMember._paginated_results(req.params, main_members)
                        resp.body = json.dumps([m.to_dict() for m in main_members.all()], default=str)

        except:
            logger.exception("Error, Failed to get Applicants for user with ID {}.".format(id))
            raise falcon.HTTPUnprocessableEntity(title="Uprocessable entity", description="Failed to get Applicants for user with ID {}.".format(id))


class MainGetAllConsultantEndpoint:
    cors = public_cors

    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def on_get(self, req, resp, id):
        try:
            with db.transaction() as session:
                try:
                    status = None
                    search_field = None
                    notice = None
                    consultant_id = None

                    if "status" in req.params:
                        status = req.params.pop("status")

                    if "search_string" in req.params:
                        search_field = req.params.pop("search_string")

                    if "notice" in req.params:
                        notice = req.params.pop("notice")

                    if "consultant_id" in req.params:
                        consultant_id = req.params.pop("consultant_id")
                        consultant = session.query(Consultant).filter(Consultant.id == consultant_id, Consultant.state == Consultant.STATE_ACTIVE).first()

                    if not consultant_id:
                        consultant = session.query(Consultant).filter(
                            Consultant.state == Consultant.STATE_ACTIVE,
                            Consultant.id == id
                        ).one()

                    parlour = session.query(Parlour).filter(
                        Parlour.id == consultant.parlour_id,
                        Parlour.state == Parlour.STATE_ACTIVE
                    ).one()

                except MultipleResultsFound as e:
                    raise falcon.HTTPBadRequest(title="Error", description="Error getting applicants")
                except NoResultFound as e:
                    raise falcon.HTTPBadRequest(title="Error", description="No results found.")

                if search_field:
                    main_members = session.query(
                        MainMember,
                        Applicant
                    ).join(Applicant, (MainMember.applicant_id==Applicant.id)).filter(
                        MainMember.state == MainMember.STATE_ACTIVE,
                        MainMember.parlour_id == parlour.id,
                        or_(
                            MainMember.first_name.ilike('%{}%'.format(search_field)),
                            MainMember.last_name.ilike('%{}%'.format(search_field)),
                            MainMember.id_number.ilike('%{}%'.format(search_field)),
                            Applicant.policy_num.ilike('%{}%'.format(search_field))
                        )
                    )

                    if not main_members:
                        resp.body = json.dumps({"total": 0, "count": 0, "offset": 0, "limit":0, "result": []})
                    else:
                        result = MainMember._paginated_search_results(req.params, main_members)
                        resp.body = json.dumps(result, default=str)
                else:
                    applicants = session.query(Applicant).filter(
                        Applicant.state == Applicant.STATE_ACTIVE,
                        Applicant.consultant_id == consultant.id
                    )

                    if status :
                        applicants = [] if status.lower() == 'lapsed' else applicants.filter(Applicant.consultant_id == consultant.id, Applicant.status == status.lower()).limit(100).all()

                    applicant_res = [(applicant, applicant.plan) for applicant in applicants]
                    if applicant_res:
                        for applicant in applicant_res:
                            if applicant[0].plan.id == applicant[1].id:
                                max_age_limit = applicant[1].member_maximum_age if applicant[1].member_maximum_age else 0
                                min_age_limit = applicant[1].member_minimum_age if applicant[1].member_minimum_age else 0
                                main_member = session.query(MainMember).filter(
                                    MainMember.state == MainMember.STATE_ACTIVE,
                                    MainMember.applicant_id == applicant[0].id
                                ).first()
                                if main_member:
                                    id_number = main_member.id_number

                                    number = int(id_number[0:2])
                                    if number == 0:
                                        number = 2000
                                    try:
                                        dob = datetime(number, int(id_number[2:4]), int(id_number[4:6]),0,0,0,0)
                                    except:
                                        dob = datetime(number, int(id_number[2:4]), 1,0,0,0,0)
                                    now = datetime.now()
                                    age = relativedelta(now, dob)

                                    years = "{}".format(age.years)

                                    if len(years) > 2:
                                        years = years[2:]

                                    if max_age_limit and int(years) > max_age_limit:
                                        main_member.age_limit_exceeded = True
                                    elif min_age_limit and int(years) < min_age_limit:
                                        main_member.age_limit_exceeded = True
                                    session.commit()
                    applicant_ids = [applicant.id for applicant in applicants]
                    if notice:
                        main_members = session.query(MainMember).filter(
                            MainMember.state == MainMember.STATE_ACTIVE,
                            MainMember.parlour_id == parlour.id,
                            MainMember.age_limit_exceeded == True,
                            MainMember.applicant_id.in_(applicant_ids)
                        ).order_by(MainMember.id.desc())
                    else:
                        main_members = session.query(MainMember).filter(
                            MainMember.state == MainMember.STATE_ACTIVE,
                            MainMember.parlour_id == parlour.id,
                            MainMember.applicant_id.in_(applicant_ids)
                        ).order_by(MainMember.id.desc())

                    if not main_members:
                        resp.body = json.dumps({})
                    else:
                        result = MainMember._paginated_results(req.params, main_members)
                        resp.body = json.dumps([m.to_dict() for m in main_members.all()], default=str)

        except:
            logger.exception("Error, Failed to get Applicants for user with ID {}.".format(id))
            raise falcon.HTTPUnprocessableEntity(title="Uprocessable entity", description="Failed to get Applicants for user with ID {}.".format(id))


class MainGetAllArchivedParlourEndpoint:
    cors = public_cors

    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def on_get(self, req, resp, id):
        try:
            with db.transaction() as session:
                try:
                    status = None
                    search_field = None
                    if "status" in req.params:
                        status = req.params.pop("status")

                    if "search_string" in req.params:
                        search_field = req.params.pop("search_string")

                    parlour = session.query(Parlour).filter(
                        Parlour.state == Parlour.STATE_ACTIVE,
                        Parlour.id == id
                    ).one()
                except MultipleResultsFound as e:
                    raise falcon.HTTPBadRequest(title="Error", description="Error getting applicants")
                except NoResultFound as e:
                    raise falcon.HTTPBadRequest(title="Error", description="Error getting applicants")

                if search_field:
                    main_members = session.query(
                        MainMember,
                        Applicant
                    ).join(Applicant, (MainMember.applicant_id==Applicant.id)).filter(
                        MainMember.state == MainMember.STATE_ARCHIVED,
                        MainMember.parlour_id == parlour.id,
                        or_(
                            MainMember.first_name.ilike('{}%'.format(search_field)),
                            MainMember.last_name.ilike('{}%'.format(search_field)),
                            MainMember.id_number.ilike('{}%'.format(search_field)),
                            Applicant.policy_num.ilike('{}%'.format(search_field))
                        )
                    ).all()

                    if not main_members:
                        resp.body = json.dumps([])

                    resp.body = json.dumps([main_member[0].to_dict() for main_member in main_members], default=str)
                else:
                    applicants = session.query(Applicant).filter(
                        Applicant.state == Applicant.STATE_ARCHIVED,
                        Applicant.parlour_id == parlour.id
                    ).order_by(Applicant.id.desc())

                    if status:
                        applicants = applicants.filter(Applicant.status == status.lower()).all()

                    applicant_ids = [applicant.id for applicant in applicants]
                    main_members = session.query(MainMember).filter(
                        or_(MainMember.state == Applicant.STATE_ARCHIVED,
                            MainMember.applicant_id.in_(applicant_ids)),
                            MainMember.parlour_id == parlour.id
                    ).all()

                    if not main_members:
                        resp.body = json.dumps([])
                    else:
                        resp.body = json.dumps([main_member.to_dict() for main_member in main_members], default=str)

        except:
            logger.exception("Error, Failed to get Applicants for user with ID {}.".format(id))
            raise falcon.HTTPUnprocessableEntity(title="Uprocessable entity", description="Failed to get Applicants for user with ID {}.".format(id))


class MainGetAllArchivedConsultantEndpoint:
    cors = public_cors

    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def on_get(self, req, resp, id):
        try:
            with db.transaction() as session:
                try:
                    status = None
                    search_field = None

                    if "status" in req.params:
                        status = req.params.pop("status")

                    if "search_string" in req.params:
                        search_field = req.params.pop("search_string")

                    consultant = session.query(Consultant).filter(
                        Consultant.state == Consultant.STATE_ACTIVE,
                        Consultant.id == id
                    ).one_or_none()
                except MultipleResultsFound as e:
                    raise falcon.HTTPBadRequest(title="Error", description="Error getting applicants")

                if search_field:
                    extended_members = session.query(ExtendedMember).filter(ExtendedMember.state == ExtendedMember.STATE_ARCHIVED).all()
                    applicant_ids = [ext.applicant_id for ext in extended_members]
                    main_members = session.query(
                        MainMember,
                        Applicant
                    ).join(Applicant, (MainMember.applicant_id==Applicant.id)).filter(
                        or_(MainMember.state == MainMember.STATE_ARCHIVED,
                        MainMember.applicant_id.in_(applicant_ids)),
                        or_(
                            MainMember.first_name.ilike('{}%'.format(search_field)),
                            MainMember.last_name.ilike('{}%'.format(search_field)),
                            MainMember.id_number.ilike('{}%'.format(search_field)),
                            Applicant.policy_num.ilike('{}%'.format(search_field))
                        ),
                        Applicant.consultant_id == consultant.id
                    ).all()

                    if not main_members:
                        resp.body = json.dumps([])

                    resp.body = json.dumps([main_member[0].to_dict() for main_member in main_members], default=str)
                else:
                    extended_members = session.query(ExtendedMember).filter(
                        or_(ExtendedMember.state == ExtendedMember.STATE_ARCHIVED,
                        ExtendedMember.is_deceased == True)
                    ).all()
                    applicant_ids = [ext.applicant_id for ext in extended_members]
                    applicants = session.query(Applicant).filter(
                        or_(Applicant.state == Applicant.STATE_ARCHIVED,
                        Applicant.id.in_(applicant_ids)),
                        Applicant.consultant_id == consultant.id
                    ).order_by(Applicant.id.desc())

                    if status:
                        applicants = applicants.filter(Applicant.status == status.lower()).all()

                    applicant_ids = [applicant.id for applicant in applicants]
                    main_members = session.query(MainMember).filter(
                        MainMember.applicant_id.in_(applicant_ids)
                    ).all()

                    if not main_members:
                        resp.body = json.dumps([])
                    else:
                        resp.body = json.dumps([main_member.to_dict() for main_member in main_members], default=str)

        except:
            logger.exception("Error, Failed to get Applicants for user with ID {}.".format(id))
            raise falcon.HTTPUnprocessableEntity(title="Uprocessable entity", description="Failed to get Applicants for user with ID {}.".format(id))


class MemberCertificateGetEndpoint:
    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def on_get(self, req, resp, id):
        try:
            with db.transaction() as session:
                # req = json.loads(req.stream.read().decode('utf-8'))

                main_meber = session.query(MainMember).filter(
                    MainMember.id == id,
                    MainMember.state != MainMember.STATE_DELETED
                ).first()
                if main_meber is None:
                    raise falcon.HTTPNotFound(title="Error", description="Main member not found")

                applicant = session.query(Applicant).filter(
                    Applicant.id == main_meber.applicant_id,
                    Applicant.state != Applicant.STATE_DELETED
                ).first()

                if applicant is None:
                    raise falcon.HTTPNotFound(title="Error", description="Applicant not found")

                with open(applicant.document, 'rb') as f:
                    resp.downloadable_as = applicant.document
                    resp.content_type = 'application/pdf'
                    resp.stream = [f.read()]
                    resp.status = falcon.HTTP_200

        except:
            logger.exception("Error, Failed to get Payment with ID {}.".format(id))
            raise falcon.HTTPUnprocessableEntity(title="Uprocessable entity", description="Failed to get Invoice with ID {}.".format(id))


class MemberPersonalDocsGetEndpoint:
    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def on_get(self, req, resp, id):
        try:
            with db.transaction() as session:

                main_meber = session.query(MainMember).filter(
                    MainMember.id == id,
                    MainMember.state == MainMember.STATE_ACTIVE
                ).first()
                if main_meber is None:
                    raise falcon.HTTPNotFound(title="Error", description="Main member not found")

                applicant = session.query(Applicant).filter(
                    Applicant.id == main_meber.applicant_id,
                    Applicant.state == Applicant.STATE_ACTIVE
                ).first()

                if applicant is None:
                    raise falcon.HTTPNotFound(title="Error", description="Applicant not found")

                with open(applicant.personal_docs, 'rb') as f:
                    resp.downloadable_as = applicant.personal_docs
                    resp.content_type = 'application/pdf'
                    resp.stream = [f.read()]
                    resp.status = falcon.HTTP_200

        except:
            logger.exception("Error, Failed to get Payment with ID {}.".format(id))
            raise falcon.HTTPUnprocessableEntity(title="Uprocessable entity", description="Failed to get Invoice with ID {}.".format(id))


class MainMemberPostFileEndpoint:

    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def on_post(self, req, resp, id):
        from base64 import b64decode
        body = json.load(req.bounded_stream)

        if not body:
            raise falcon.HTTPBadRequest(title='400 Bad Request', description='Body is empty or malformed.')

        pdf = body['pdf']

        with db.transaction() as session:
            try:
                try:
                    main_member = session.query(MainMember).filter(
                        MainMember.id == id,
                        MainMember.state == MainMember.STATE_ACTIVE
                    ).one_or_none()

                except NoResultFound:
                    raise falcon.HTTPBadRequest(title="Error", description="No results found.")
                except MultipleResultsFound:
                    raise falcon.HTTPBadRequest(title="Error", description="Multiple results found.")

                applicant = session.query(Applicant).filter(
                    Applicant.id == main_member.applicant_id,
                    Applicant.state == Applicant.STATE_ACTIVE).first()

                if not applicant:
                    raise falcon.HTTPBadRequest(title="Error", description="No results found.")

                filename = "{uuid}.{ext}".format(uuid=uuid.uuid4(), ext='pdf')

                os.chdir('./assets/uploads/personal_docs')
                pdf_path = os.path.join(os.getcwd(), filename)
                with open(pdf_path, "wb") as pdf_file:
                    bytes = b64decode(pdf)
                    pdf_file.write(bytes)

                applicant.personal_docs = '{}/{}'.format(os.getcwd(), filename)
                os.chdir('../../..')
                resp.status = falcon.HTTP_200
                resp.location = filename

                resp.body = json.dumps({})

            except Exception as e:
                logger.exception(
                    "Error, experienced error while creating Applicant.")
                raise e


class MainMemberPostEndpoint:

    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def get_date(self, input_date):
        if input_date:
            return input_date.replace('T', " ")[:10]
        return None

    def on_post(self, req, resp, id):
        
        req = json.load(req.bounded_stream)

        with db.transaction() as session:
            parlour = session.query(Parlour).filter(
                Parlour.id == req["parlour_id"],
                Parlour.state == Parlour.STATE_ACTIVE).first()

            if not parlour:
                raise falcon.HTTPBadRequest("Parlour does not exist.")
            if not req.get("id_number"):
                raise falcon.HTTPBadRequest(title="Error", description="Missing id_number field.")
            if not req.get("first_name"):
                raise falcon.HTTPBadRequest(title="Error", description="Missing first name field.")
            if not req.get("last_name"):
                raise falcon.HTTPBadRequest(title="Error", description="Missing last name field.")
            if not req.get("date_joined"):
                    raise falcon.HTTPNotFound(title="Error", description="Date joined is a required field.")
            if not req.get("contact"):
                    raise falcon.HTTPNotFound(title="Error", description="Contact number is a required field.")

            consultant = session.query(Consultant).get(id)

            if not consultant:
                raise falcon.HTTPBadRequest(title="Error", description="Consultant does not exist.")

            plan_id = req.get("plan_id")

            plan = session.query(Plan).filter(
                Plan.id == plan_id,
                Plan.state == Plan.STATE_ACTIVE).one_or_none()

            if not plan:
                raise falcon.HTTPBadRequest(title="Error", description="Plan does not exist.")

            applicant_req = req.get("applicant")

            if not applicant_req.get("policy_num"):
                raise falcon.HTTPBadRequest(title="Error", description="Missing policy number field.")

            applicants = session.query(Applicant).filter(Applicant.plan_id == plan.id).all()
            applicant_ids = [applicant.id for applicant in applicants]

            id_number = session.query(MainMember).filter(
                MainMember.id_number == req.get("id_number"),
                MainMember.state.in_((MainMember.STATE_ACTIVE, MainMember.STATE_ARCHIVED)),
                MainMember.applicant_id.in_(applicant_ids)
            ).first()

            if not id_number:
                id_number = session.query(ExtendedMember).filter(
                    ExtendedMember.id_number == req.get("id_number"),
                    ExtendedMember.state.in_((ExtendedMember.STATE_ACTIVE, ExtendedMember.STATE_ARCHIVED)),
                    ExtendedMember.applicant_id.in_(applicant_ids)
                ).first()

            if id_number:
                raise falcon.HTTPBadRequest(title="Error", description="ID number already exists for either main member or extended member.")

            try:
                applicant = Applicant(
                    policy_num=applicant_req.get("policy_num"),
                    address=applicant_req.get("address"),
                    status='unpaid',
                    plan_id=plan.id,
                    consultant_id=consultant.id,
                    parlour_id=parlour.id,
                    old_url=False,
                    date=datetime.now(),
                    state=Applicant.STATE_ACTIVE,
                    modified_at=datetime.now(),
                    created_at=datetime.now()
                )

                applicant.save(session)
                date_joined = self.get_date(req.get("date_joined"))
                date_of_birth = self.get_date(req.get("date_of_birth"))
                main_member = MainMember(
                    first_name = req.get("first_name"),
                    last_name = req.get("last_name"),
                    id_number = req.get("id_number"),
                    contact = req.get("contact"),
                    date_of_birth = date_of_birth,
                    parlour_id = parlour.id,
                    waiting_period = req.get("waiting_period", 0),
                    date_joined = date_joined,
                    state=MainMember.STATE_ACTIVE,
                    applicant_id = applicant.id,
                    modified_at = datetime.now(),
                    created_at = datetime.now()
                )

                min_age_limit = plan.member_minimum_age
                max_age_limit = plan.member_maximum_age

                id_number = main_member.id_number
                if int(id_number[0:2]) > 21:
                    number = '19{}'.format(id_number[0:2])
                else:
                    number = '20{}'.format(id_number[0:2])
                dob = parse('{}-{}-{}'.format(number, id_number[2:4], id_number[4:6]))

                now = datetime.now()

                age = relativedelta(now, dob)

                years = "{}".format(age.years)
                try:
                    if max_age_limit and len(years) > 2 and int(years[2:4]) > max_age_limit:
                        main_member.age_limit_exceeded = True
                    elif max_age_limit and int(years) > max_age_limit:
                        main_member.age_limit_exceeded = True
                except:
                    pass

                try:
                    if min_age_limit and len(years) > 2 and int(years[2:4]) < min_age_limit:
                        main_member.age_limit_exceeded = True
                    elif min_age_limit and int(years) < min_age_limit:
                        main_member.age_limit_exceeded = True
                except:
                    pass
                main_member.save(session)

                applicant = update_certificate(applicant)

                resp.body = json.dumps(main_member.to_dict(), default=str)
            except Exception as e:
                logger.exception(
                    "Error, experienced error while creating Applicant.")
                raise e


class MainMemberBulkPostEndpoint:

    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def get_date(self, input_date):
        if input_date:
            return input_date.replace('T', " ")[:10]
        return None

    @staticmethod
    def format_csv(csv_data):
        csv_data = csv_data.splitlines()
        result = []

        count = 1
        for row in csv.reader(csv_data):
            if count > 1:
                result.append(row)
            count +=1

        return result

    def on_post(self, req, resp, id):
        import io
        from base64 import b64decode
        import openpyxl
        rest_dict = json.load(req.bounded_stream)
        error_data = []
        prev_applicant = None
        is_main_member = False
        plan_id = rest_dict.pop('plan')
        csv_data = rest_dict.pop('csv')
        filename = uuid.uuid4()

        os.chdir('./assets/uploads/spreadsheets')
        with open('{}.xlsx'.format(filename), "wb") as file:
            file.write(b64decode(csv_data))

        wookbook = openpyxl.load_workbook('{}.xlsx'.format(filename))

        # Define variable to read the active sheet:
        worksheet = wookbook.active
        rows = []
        # Iterate the loop to read the cell values
        for i in range(0, worksheet.max_row):
            if not i == 0:
                rows.append([])
                for col in worksheet.iter_cols(1, worksheet.max_column):
                    rows[i-1].append(col[i].value)

        os.remove('{}.xlsx'.format(filename))
        os.chdir('../../..')

        for data in rows:

            with db.transaction() as session:
                id_check = '{}'.format(data[2])
                if not any(data):
                    continue
                try:
                    if data[8] or data[9]:
                        is_main_member = False
                        if not all([data[8], data[9]]):
                            error_data.append({'data': data, 'error': "Extended member type and relation to main member are requires fieds."})
                            continue
                        if not prev_applicant:
                            error_data.append({'data': data, 'error': "Main member to extended member has an issue."})
                            continue
                        error_data = bulk_insert_extended_members(data, error_data, prev_applicant, session)
                        continue
                    else:
                        is_main_member = True
                except IndexError:
                    pass
                if not id_check:
                    prev_applicant = None if is_main_member else prev_applicant
                    error_data.append({'data': data, 'error': "Missing id_number field."})
                    continue

                za_validation = SouthAfricanIdentityValidate(id_check)
                valid = za_validation.validate()
                if not valid:
                    prev_applicant = None if is_main_member else prev_applicant
                    error_data.append({'data': data, 'error': "Incorrect id_number entered."})
                    continue

                if not data[0]:
                    prev_applicant = None if is_main_member else prev_applicant
                    error_data.append({'data': data, 'error': "Missing first name field."})
                    continue
                if not data[1]:
                    prev_applicant = None if is_main_member else prev_applicant
                    error_data.append({'data': data, 'error': "Missing last name field."})
                    continue
                if not data[4]:
                    prev_applicant = None if is_main_member else prev_applicant
                    error_data.append({'data': data, 'error': "Date joined is a required field."})
                    continue
                if not data[3]:
                    prev_applicant = None if is_main_member else prev_applicant
                    error_data.append({'data': data, 'error': "Contact number is a required field."})
                    continue

                consultant = session.query(Consultant).get(id)

                if not consultant:
                    prev_applicant = None if is_main_member else prev_applicant
                    error_data.append({'data': data, 'error': "Consultant does not exist."})
                    continue

                plan = session.query(Plan).get(plan_id)

                if not plan:
                    prev_applicant = None if is_main_member else prev_applicant
                    error_data.append({'data': data, 'error': "Plan not found."})
                    continue

                parlour = session.query(Parlour).filter(
                    Parlour.id == plan.parlour_id,
                    Parlour.state == Parlour.STATE_ACTIVE).first()

                if not parlour:
                    prev_applicant = None if is_main_member else prev_applicant
                    error_data.append({'data': data, 'error': "Parlour does not exist."})
                    continue

                if not data[7]:
                    prev_applicant = None if is_main_member else prev_applicant
                    error_data.append({'data': data, 'error': 'Missing policy number'})
                    continue

                id_number = session.query(MainMember).filter(
                    MainMember.id_number == id_check,
                    MainMember.state.in_((MainMember.STATE_ACTIVE, MainMember.STATE_ARCHIVED)),
                    MainMember.parlour_id == parlour.id
                ).first()

                if not id_number:
                    applicants = session.query(Applicant).filter(Applicant.plan_id == plan.id).all()
                    applicant_ids = [applicant.id for applicant in applicants]
                    id_number = session.query(ExtendedMember).filter(
                        ExtendedMember.id_number == id_check,
                        ExtendedMember.state.in_((ExtendedMember.STATE_ACTIVE, ExtendedMember.STATE_ARCHIVED)),
                        ExtendedMember.applicant_id.in_(applicant_ids)
                    ).first()

                if id_number:
                    prev_applicant = None if is_main_member else prev_applicant
                    if isinstance(id_number,MainMember):
                        prev_applicant = id_number.applicant_id
                    error_data.append({'data': data, 'error': 'ID number already exists'})
                    continue

                applicant = Applicant(
                    policy_num=data[7],
                    address=data[6],
                    status='unpaid',
                    plan_id=plan.id,
                    consultant_id=consultant.id,
                    parlour_id=parlour.id,
                    old_url=False,
                    date=datetime.now(),
                    state=Applicant.STATE_ACTIVE,
                    modified_at=datetime.now(),
                    created_at=datetime.now()
                )

                applicant.save(session)
                prev_applicant = applicant.id

                date_joined = data[4]
                main_member = MainMember(
                    first_name = data[0],
                    last_name = data[1],
                    id_number = id_check,
                    contact = data[3] if len(str(data[3])) == 10 else '0{}'.format(data[3]),
                    parlour_id = parlour.id,
                    date_joined = date_joined,
                    waiting_period = data[5] if data[5] else 0,
                    state=MainMember.STATE_ACTIVE,
                    applicant_id = applicant.id,
                    modified_at = datetime.now(),
                    created_at = datetime.now()
                )

                min_age_limit = plan.member_minimum_age
                max_age_limit = plan.member_maximum_age

                id_number = main_member.id_number

                if int(id_number[0:2]) > (int(str(datetime.now().year)[2:]) - 16):
                    number = '19{}'.format(id_number[0:2])
                else:
                    number = '20{}'.format(id_number[0:2])

                try:
                    dob = parse('{}/{}/{}'.format(number, id_number[2:4], id_number[4:6]))
                except ValueError:
                    prev_applicant = None if is_main_member else prev_applicant
                    error_data.append({'data': data, 'error': 'Incorrect date formt on date of birth'})
                    continue

                now = datetime.now()

                age = relativedelta(now, dob)

                years = "{}".format(age.years)
                try:
                    if max_age_limit and len(years) > 2 and int(years[2:4]) > max_age_limit:
                        main_member.age_limit_exceeded = True
                    elif max_age_limit and int(years) > max_age_limit:
                        main_member.age_limit_exceeded = True
                except:
                    pass

                try:
                    if min_age_limit and len(years) > 2 and int(years[2:4]) < min_age_limit:
                        main_member.age_limit_exceeded = True
                    elif min_age_limit and int(years) < min_age_limit:
                        main_member.age_limit_exceeded = True
                except:
                    pass
                main_member.save(session)

                applicant = update_certificate(applicant)

        resp.body = json.dumps(error_data, default=str)


class MainMemberCheckAgeLimitEndpoint:
    cors = public_cors

    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def get_date_of_birth(self, date_of_birth=None, id_number=None):
        current_year = datetime.now().year
        year_string = str(current_year)[2:]
        century = 19
        if date_of_birth:
            return date_of_birth.replace('T', " ")[:10]
        if id_number:
            if 0 <= int(id_number[:2]) <= int(year_string):
                century = 20
            return '{}{}-{}-{}'.format(century,id_number[:2], id_number[2:4], id_number[4:-6])[:10]

    def on_get(self, req, resp, id):

        with db.no_transaction() as session:
            age_limit_exceeded = False
            id_number_param = None
            date_of_birth = None
            plan = None
            max_age_limit = None
            min_age_limit = None

            try:
                id_number_param = req.params.pop("id_number")
            except:
                raise falcon.HTTPBadRequest(title="Error", description="Missing id_number field.")

            plan = session.query(Plan).get(id)

            parlour = plan.parlour
            if not plan:
                raise falcon.HTTPBadRequest(title="Plan not found", description="Plan does not exist.")
 
            is_ID_number = session.query(MainMember).filter(
                MainMember.id_number == id_number_param,
                MainMember.state.in_((MainMember.STATE_ACTIVE, MainMember.STATE_ARCHIVED)),
                MainMember.parlour_id == parlour.id
            ).first()

            if not is_ID_number:
                applicants = session.query(Applicant).filter(Applicant.parlour_id == parlour.id).all()
                applicant_ids = [applicant.id for applicant in applicants]
                is_ID_number = session.query(ExtendedMember).filter(
                    ExtendedMember.id_number == id_number_param,
                    ExtendedMember.state.in_((ExtendedMember.STATE_ACTIVE, ExtendedMember.STATE_ARCHIVED)),
                    ExtendedMember.applicant_id.in_(applicant_ids)
                ).first()

            min_age_limit = plan.member_minimum_age
            max_age_limit = plan.member_maximum_age

            if int(id_number_param[0:2]) > 21:
                number = '19{}'.format(id_number_param[0:2])
            else:
                number = '20{}'.format(id_number_param[0:2])
            try:
                date_of_birth = '{}-{}-{}'.format(number, id_number_param[2:4], id_number_param[4:6])
                dob = parse(self.get_date_of_birth(date_of_birth, id_number_param)).date()
            except:
                raise falcon.HTTPBadRequest(title="Plan not found", description="Encountered error while formating date. Make sure you've entered a valid date.")
            now = datetime.now().date()

            age = relativedelta(now, dob)

            years = "{}".format(age.years)

            if max_age_limit:
                if len(years) > 2 and int(years[2:4]) > max_age_limit:
                    age_limit_exceeded = True
                elif int(years) > max_age_limit:
                    age_limit_exceeded = True

            if min_age_limit:
                if len(years) > 2 and int(years[2:4]) < min_age_limit:
                    age_limit_exceeded = True
                elif int(years) < min_age_limit:
                    age_limit_exceeded = True

            if age_limit_exceeded or is_ID_number:
                id_exists = True if is_ID_number else False
                resp.body = json.dumps({'age_limit_exceeded': age_limit_exceeded, 'id_number_exists': id_exists })
            else:
                resp.body = json.dumps({'result': 'OK!'})

class MainMemberPutEndpoint:

    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def get_date(self, input_date):
        if input_date:
            return input_date.replace('T', " ")[:10]
        return None

    def on_put(self, req, resp, id):
        req = json.load(req.bounded_stream)

        with db.transaction() as session:

            try:
                parlour_id = req.get("parlour_id")

                parlour = session.query(Parlour).filter(
                    Parlour.id == parlour_id).one()
            except MultipleResultsFound:
                raise falcon.HTTPBadRequest(title="Error", description="Bad Request")
            except NoResultFound:
                raise falcon.HTTPNotFound(title="Not Found", description="Parlour not found")

            if not req.get("id_number"):
                raise falcon.HTTPBadRequest(title="Error", description="Missing id_number field.")
            if not req.get("first_name"):
                raise falcon.HTTPBadRequest(title="Error", description="Missing first name field.")
            if not req.get("last_name"):
                raise falcon.HTTPBadRequest(title="Error", description="Missing last name field.")


            applicant_req = req.get("applicant")

            plan_id = req.get("plan_id")

            plan = session.query(Plan).filter(
                Plan.id == plan_id,
                Plan.state == Plan.STATE_ACTIVE).one_or_none()

            applicant = session.query(Applicant).filter(
                Applicant.id == applicant_req.get("id"),
                Applicant.state == Applicant.STATE_ACTIVE).first()

            if not applicant:
                raise falcon.HTTPNotFound(title="Applicant not found", description="Could not find Applicant with given ID.")

            if plan:
                applicant.plan_id = plan.id
            applicant.policy_num = applicant_req.get("policy_num")
            applicant.address = applicant_req.get("address")

            main_member = session.query(MainMember).filter(
                MainMember.id == id,
                MainMember.parlour_id == parlour.id,
                MainMember.state == MainMember.STATE_ACTIVE).first()

            if not main_member:
                raise falcon.HTTPNotFound(title="Main member not found", description="Could not find Applicant with given ID.")

            applicants = session.query(Applicant).filter(Applicant.plan_id == plan.id).all()
            applicant_ids = [applicant.id for applicant in applicants]
            id_number = session.query(MainMember).filter(
                MainMember.id_number == req.get("id_number"),
                MainMember.applicant_id.in_(applicant_ids),
                MainMember.id != main_member.id,
                MainMember.state.in_((MainMember.STATE_ACTIVE, MainMember.STATE_ARCHIVED))
            ).first()

            if not id_number:
                id_number = session.query(ExtendedMember).filter(
                    ExtendedMember.id_number == req.get("id_number"),
                    ExtendedMember.state.in_((ExtendedMember.STATE_ACTIVE, ExtendedMember.STATE_ARCHIVED)),
                    ExtendedMember.applicant_id.in_(applicant_ids)
                ).first()

            if id_number:
                raise falcon.HTTPBadRequest(title="Error", description="ID number already exists for either main member or extended member.")

            try:
                main_member.first_name = req.get("first_name")
                main_member.last_name = req.get("last_name")
                main_member.id_number = req.get("id_number")
                main_member.contact = req.get("contact")
                main_member.date_of_birth = req.get("date_of_birth")
                main_member.waiting_period = req.get("waiting_period", 0)
                main_member.parlour_id = parlour.id
                main_member.applicant_id = applicant.id
                if req.get("is_deceased"):
                    main_member.is_deceased = req.get("is_deceased")
                    main_member.state = MainMember.STATE_ARCHIVED
                    applicant.state = MainMember.STATE_ARCHIVED
                    update_deceased_extended_members(session, main_member)

                main_member.save(session)
                applicant = update_certificate(applicant)
                resp.body = json.dumps(main_member.to_dict(), default=str)
            except:
                logger.exception(
                    "Error, experienced error while creating Applicant.")
                raise falcon.HTTPBadRequest(
                    "Processing Failed. experienced error while creating Applicant.")


class MainMemberPutAgeLimitExceptionEndpoint:

    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def on_put(self, req, resp, id):
        req = json.load(req.bounded_stream)
        try:
            with db.transaction() as session:

                main_member = session.query(MainMember).filter(
                    MainMember.id == id,
                    MainMember.state == MainMember.STATE_ACTIVE).first()

                if not main_member:
                    raise falcon.HTTPNotFound(title="Main member not found", description="Could not find Applicant with given ID.")

                main_member.age_limit_exception = req.get("age_limit_exception")

                main_member.save(session)

                resp.body = json.dumps(main_member.to_dict(), default=str)
        except:
            logger.exception(
                "Error, experienced error while creating Applicant.")
            raise falcon.HTTPBadRequest(
                "Processing Failed. experienced error while creating Applicant.")


class MainMemberRestorePutEndpoint:

    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def on_put(self, req, resp, id):

        req = json.load(req.bounded_stream)

        try:
            with db.transaction() as session:
                try:
                    parlour = session.query(Parlour).filter(
                        Parlour.id == req.get("parlour")['id']).one()
                except MultipleResultsFound:
                    raise falcon.HTTPBadRequest(title="Error", description="Bad Request")
                except NoResultFound:
                    raise falcon.HTTPNotFound(title="Not Found", description="Parlour not found")

                main_member = session.query(MainMember).filter(
                    MainMember.id == id,
                    MainMember.parlour_id == parlour.id).first()

                if not main_member:
                    raise falcon.HTTPNotFound(title="Main member not found", description="Could not find Applicant with given ID.")

                main_member.state = MainMember.STATE_ACTIVE
                main_member.waiting_period = req.get('waiting_period')
                applicant = session.query(Applicant).get(main_member.applicant_id)
                applicant.state = Applicant.STATE_ACTIVE

                main_member.save(session)
                applicant = update_certificate(applicant)
                resp.body = json.dumps(main_member.to_dict(), default=str)
        except:
            logger.exception(
                "Error, experienced error while creating Applicant.")
            raise falcon.HTTPBadRequest(
                "Processing Failed. experienced error while creating Applicant.")


class MainMemberDeleteEndpoint:

    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def on_delete(self, req, resp, id):
        try:
            with db.transaction() as session:
                try:
                    main_member = session.query(MainMember).get(id)
                    applicant = main_member.applicant
                except MultipleResultsFound:
                    raise falcon.HTTPBadRequest(title="Error", description="Bad Request")
                except NoResultFound:
                    raise falcon.HTTPNotFound(title="Not Found", description="Member not found")

                if main_member.is_deleted:
                    falcon.HTTPNotFound(title="Not Found", description="Member does not exist.")

                main_member.delete(session)
                applicant.delete(session)
                resp.body = json.dumps({})
        except:
            logger.exception("Error, Failed to delete Applicant with ID {}.".format(id))
            raise falcon.HTTPBadRequest(title="Bad Request", description="Failed to delete Applicant with ID {}.".format(id))


class MainMemberArchiveEndpoint:

    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def on_delete(self, req, resp, id):
        try:
            with db.transaction() as session:
                try:
                    main_member = session.query(MainMember).get(id)
                except MultipleResultsFound:
                    raise falcon.HTTPBadRequest(title="Error", description="Bad Request")
                except NoResultFound:
                    raise falcon.HTTPNotFound(title="Not Found", description="Member not found")

                if main_member.is_archived:
                    falcon.HTTPNotFound(title="Not Found", description="Member does not exist.")

                main_member.archive(session)
                resp.body = json.dumps({})
        except:
            logger.exception("Error, Failed to delete Applicant with ID {}.".format(id))
            raise falcon.HTTPBadRequest(title="Bad Request", description="Failed to delete Applicant with ID {}.".format(id))


class MainMemberDownloadCSVGetEndpoint:

    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def on_get(self, req, resp, id):

        try:
            with db.transaction() as session:
                parlour = session.query(Parlour).filter(
                    Parlour.id == id,
                    Parlour.state == Parlour.STATE_ACTIVE).first()

                if not parlour:
                    raise falcon.HTTPBadRequest("Parlour does not exist.")

                main_members = session.query(MainMember).join(
                    Applicant, (MainMember.applicant_id==Applicant.id)
                ).join(Plan, (Applicant.plan_id==Plan.id)).filter(MainMember.parlour_id==parlour.id).all()

                if not main_members:
                    resp.body = json.dumps([])
                else:
                    resp.body = json.dumps([main_member.to_dict() for main_member in main_members], default=str)
        except:
            logger.exception(
                "Error, experienced error while creating Applicant.")
            raise falcon.HTTPBadRequest(
                "Processing Failed. experienced error while creating Applicant.")


    def csv_dump(self, session, parlour_id):

        folder = os.path.dirname(__file__)

        today = datetime.today()

        logger.info('Create CSV file...')
        filename = os.path.join(folder, 'applicants_{}.xlsx'.format(today))

        with open(filename, 'w') as f:
            writer = csv.writer(f)
            writer.writerow([
                'First Name',
                'Last Name',
                'ID Number',
                'Contact Number',
                'Policy Number',
                'Date',
                'Status',
                'Cover',
                'Premium'
            ])

            main_members = session.query(MainMember).join(
                Applicant, (MainMember.applicant_id==Applicant.id)
            ).join(Plan, (Applicant.plan_id==Plan.id)).filter(MainMember.parlour_id==parlour_id).all()

            for main_member in main_members:
                member_dict = main_member.to_dict()
                writer.writerow([
                    member_dict.get("first_name"),
                    member_dict.get("last_name"),
                    member_dict.get("id_number"),
                    member_dict.get("contact"),
                    member_dict.get("applicant").get("policy_num"),
                    member_dict.get("date_joined"),
                    member_dict.get("applicant").get("status"),
                    member_dict.get("applicant").get("cover"),
                    member_dict.get("applicant").get("premium")
                ])


class ApplicantExportToExcelEndpoint:
    cors = public_cors
    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def on_get(self, req, resp, id):
        try:
            with db.transaction() as session:
                try:
                    status = None
                    permission = None
                    parlour = None
                    consultant = None

                    if "status" in req.params:
                        status = req.params.pop("status")

                    if "permission" in req.params:
                        permission = req.params.pop("permission")

                    if permission.lower() == 'consultant':
                        consultant = session.query(Consultant).filter(
                            Consultant.state == Consultant.STATE_ACTIVE,
                            Consultant.id == id
                        ).one_or_none()

                    if permission.lower() == 'parlour':
                        parlour = session.query(Parlour).filter(
                            Parlour.state == Parlour.STATE_ACTIVE,
                            Parlour.id == id
                        ).one_or_none()
                except MultipleResultsFound as e:
                    raise falcon.HTTPBadRequest(title="Error", description="Error getting applicants")

                if consultant:
                    applicants = session.query(Applicant).filter(
                        Applicant.state == Applicant.STATE_ACTIVE,
                        Applicant.consultant_id == consultant.id
                    ).order_by(Applicant.id.desc())

                if parlour:
                    applicants = session.query(Applicant).filter(
                        Applicant.state == Applicant.STATE_ACTIVE,
                        Applicant.parlour_id == parlour.id
                    ).order_by(Applicant.id.desc())

                if status:
                    applicants = applicants.filter(Applicant.status == status.lower()).all()

                applicant_ids = [applicant.id for applicant in applicants]

                if not applicant_ids:
                    raise falcon.HTTPBadRequest(title="Error", description="No Applicants available")

                main_members = session.query(MainMember).filter(
                    MainMember.state == MainMember.STATE_ACTIVE,
                    MainMember.applicant_id.in_(applicant_ids)
                ).all()
                results = []

                for main in main_members:
                    d = main.to_short_dict()
                    results.append(d)

                    extended_members = session.query(ExtendedMember).filter(
                        ExtendedMember.state == ExtendedMember.STATE_ACTIVE,
                        ExtendedMember.applicant_id == main.applicant.id
                    ).all()
                    for ex in extended_members:
                       e = ex.to_short_dict()
                       results.append(e)

                if results:
                    data = []
                    for res in results:
                        applicant = res.get('applicant')

                        plan = applicant.get('plan')
                        underwriter = float(plan.get('underwriter_premium')) if plan.get('underwriter_premium') else None
                        data.append({
                            'First Name': res.get('first_name'),
                            'Last Name': res.get('last_name'),
                            'ID Number': res.get('id_number') if res.get('id_number') else res.get('date_of_birth'),
                            'Policy Number': applicant.get("policy_num"),
                            'Contact Number': res.get('contact') if res.get('contact') else res.get('number'),
                            'Date Joined': res.get('date_joined') if res.get('date_joined') else None,
                            'Status': applicant.get('status') if res.get else None,
                            'Premium': None if res.get('relation_to_main_member') else float(plan.get('premium')),
                            'Underwriter': None if res.get('relation_to_main_member') else underwriter,
                            'Relation to Main Member': ExtendedMember.relation_to_text.get(res.get('relation_to_main_member')) if res.get('relation_to_main_member') else None,
                            })

                    df = pd.DataFrame(data)
                    filename = '{}_{}'.format(consultant.first_name, consultant.last_name) if consultant else parlour.parlourname
                    writer = pd.ExcelWriter('{}.xlsx'.format(filename), engine='xlsxwriter')
                    df.to_excel(writer, sheet_name='Sheet1', index=False)
                    os.chdir('./assets/uploads/spreadsheets')
                    path = os.getcwd()
                    writer.save()
                    os.chdir('../../..')

                    with open('{}/{}.xlsx'.format(path, filename), 'rb') as f:
                        resp.downloadable_as = '{}.xls'.format(filename)
                        resp.content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        resp.stream = [f.read()]
                        resp.status = falcon.HTTP_200

        except Exception as e:
            logger.exception("Error, Failed to get Applicants for user with ID {}.".format(id))
            raise e


class DownloadFailedMembers:
    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def on_post(self, req, resp):
        rest_dict = json.load(req.bounded_stream)
        data = []

        for res in rest_dict:
            data.append({
                'First Name': res.get('first_name'),
                'Last Name': res.get('last_name'),
                'ID Number': res.get('id_number'),
                'Contact Number': res.get('contact'),
                'Date Joined': res.get('date_joined'),
                'Waiting Period': res.get('waiting_period', 0),
                'Physical Address': res.get('physical_address'),
                'Policy': res.get('policy_num'),
                'Type Member': res.get('type_member'),
                'Relation to Main Member': res.get('relation'),
                'Reason': res.get('reason')
                })

        df = pd.DataFrame(data)
        filename = uuid.uuid4()
        writer = pd.ExcelWriter('{}.xlsx'.format(filename), engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        os.chdir('./assets/uploads/spreadsheets')
        path = os.getcwd()
        writer.save()
        os.chdir('../../..')

        resp.body = json.dumps({'filename': filename}, default=str)


class FailedMembersExcel:
    cors = public_cors
    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def on_get(self, req, resp):
        filename = req.params.pop("filename")
        os.chdir('./assets/uploads/spreadsheets')
        path = os.getcwd()
        os.chdir('../../..')
        with open('{}/{}.xlsx'.format(path, filename), 'rb') as f:
            resp.downloadable_as = '{}.xls'.format(filename)
            resp.content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            resp.stream = [f.read()]
            resp.status = falcon.HTTP_200


class SMSService:
    def __init__(self, secure=False, basic_secure=False):
        self.secure = secure
        self.basic_secure = basic_secure

    def is_basic_secure(self):
        return self.basic_secure

    def is_not_secure(self):
        return not self.secure

    def on_post(self, req, resp):
        import requests
        rest_dict = json.load(req.bounded_stream)
        with db.transaction() as session:
            status = None
            search_field = None
            contacts = []
            consultant = None

            message = rest_dict.get("message")

            if not message:
                raise falcon.HTTPBadRequest(title="Missing Field", description="Message cannot be empty when sending an sms.")

            if rest_dict["to"]:
                cons = rest_dict.get("to").split(',')
                for contact in cons:
                    contacts.append(localize_contact(contact.strip()))

            if rest_dict.get('state'):
                status = rest_dict.get("state")

            if rest_dict.get("search_string"):
                search_field = rest_dict.get("search_string")

            parlour = session.query(Parlour).filter(
                Parlour.id == rest_dict.get("parlour_id"),
                Parlour.state == Parlour.STATE_ACTIVE).first()

            if not parlour:
                raise falcon.HTTPBadRequest(title="Error", description="Error getting applicants")

            if rest_dict.get('consultant_id'):
                consultant = session.query(Consultant).filter(
                    Consultant.id == rest_dict.get('consultant_id'),
                    Consultant.state == Consultant.STATE_ACTIVE
                ).first()

            if search_field:
                main_members = session.query(
                    MainMember,
                    Applicant
                ).join(Applicant, (MainMember.applicant_id==Applicant.id)).filter(
                    MainMember.state == MainMember.STATE_ACTIVE,
                    Applicant.parlour_id == parlour.id,
                    or_(
                        MainMember.first_name.ilike('{}%'.format(search_field)),
                        MainMember.first_name.ilike('{}%'.format(search_field)),
                        MainMember.id_number.ilike('{}%'.format(search_field)),
                        Applicant.policy_num.ilike('{}%'.format(search_field))
                    )
                )
                if consultant:
                    main_members = main_members.filter(Applicant.consultant_id == consultant.id)

            if status:
                applicants = session.query(Applicant).filter(
                    Applicant.status == status,
                    Applicant.parlour_id== parlour.id
                )
                if consultant:
                    applicants = applicants.filter(Applicant.consultant_id == consultant.id)
                applicant_ids = [applicant.id for applicant in applicants.all()]
                main_members = session.query(MainMember).filter(
                    MainMember.state == MainMember.STATE_ACTIVE,
                    MainMember.applicant_id.in_(applicant_ids)
                )

            if not status and not search_field:
                applicants = session.query(Applicant).filter(
                    Applicant.state == Applicant.STATE_ACTIVE,
                    Applicant.parlour_id== parlour.id
                )
                if consultant:
                    applicants = applicants.filter(Applicant.consultant_id == consultant.id)
                applicant_ids = [applicant.id for applicant in applicants.all()]
                main_members = session.query(MainMember).filter(
                    MainMember.state == MainMember.STATE_ACTIVE,
                    MainMember.applicant_id.in_(applicant_ids)
                )

            if rest_dict['start_date']:
                start_date = rest_dict['start_date']
                main_members = main_members.filter(
                    MainMember.created_at >= start_date
                )

            if rest_dict['end_date']:
                end_date = rest_dict['end_date']
                main_members = main_members.filter(
                    MainMember.created_at <= end_date
                )

            if not contacts:
                contacts = [m.localize_contact() for m in main_members.all()]

            if parlour.number_of_sms < len(contacts):
                raise falcon.HTTPBadRequest(title="Error", description="You need more smses to use this service.")

            headers = {"Content-Type": "application/json; charset=utf-8", "Authorization": conf.SMS_AUTH_TOKEN}
            response = requests.post('https://api.bulksms.com/v1/messages', headers=headers, json={'from': conf.SMS_FROM_NUMBER, 'to': contacts,  'body': '{}: {}'.format(parlour.parlourname.title(), message)})
            parlour.number_of_sms = parlour.number_of_sms - len(contacts) if parlour.number_of_sms > len(contacts) else 0

            result = {'status_code': response.status_code, 'parlour': parlour.to_dict()}
            resp.body = json.dumps(result, default=str)


def update_deceased_extended_members(session, main_member):
    extended_members = session.query(ExtendedMember).filter(ExtendedMember.applicant_id == main_member.applicant_id).all()
    main_member.save(session)
    for x in extended_members:
        x.is_main_member_deceased = True
    session.commit()


def update_deceased_member(session, member):
    applicant = member.applicant
    extended_members = session.query(ExtendedMember).filter(ExtendedMember.applicant_id == member.applicant_id).all()
    spouse = extended_members.pop()

    new_applicant = Applicant(
        policy_num=applicant.policy_num,
        address=applicant.address,
        status='unpaid',
        plan_id=applicant.plan_id,
        consultant_id=applicant.consultant_id,
        parlour_id=applicant.parlour_id,
        old_url=False,
        date=datetime.now(),
        state=Applicant.STATE_ACTIVE,
        modified_at=datetime.now(),
        created_at=datetime.now()
    )

    new_applicant.save(session)
    date_joined = member.date_joined
    main_member = MainMember(
        first_name=spouse.first_name,
        last_name=spouse.last_name,
        id_number=spouse.id_number,
        contact='0796579128',
        date_of_birth=spouse.date_of_birth,
        parlour_id=member.parlour_id,
        date_joined=date_joined,
        state=MainMember.STATE_ACTIVE,
        applicant_id=new_applicant.id,
        modified_at=datetime.now(),
        created_at=datetime.now()
    )

    main_member.save(session)
    for x in extended_members:
        x.applicant_id = new_applicant.id
    session.commit()
